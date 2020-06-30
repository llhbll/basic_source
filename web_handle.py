999 브라우저 띄우지 않고 data 가져오기
import requests # import urllib.request
from bs4 import BeautifulSoup

base_url = 'http://movie.daum.net/moviedb/grade?movieId=2725&type=netizen&page={}'

for n in range(77):
    url = base_url.format(n + 1)
    webpage = requests.get(url).text # webpage = urllib.request.urlopen(url)와 같음
    source = BeautifulSoup(webpage, 'html5lib')
    reviews = source.find_all('p', {'class': 'desc_review'})

    for review in reviews:
        print(review.get_text().strip())


999 포탈 자동로그인 IE는 에러가 계속나서 전자결재는 OPEN안함
import os
import sys
import time

from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys

#pyinstaller --noconsole --onefile --add-binary "chromedriver.exe";"." gmail_auto.py
if getattr(sys, 'frozen', False):
    chromedriver_path = os.path.join(sys._MEIPASS, "chromedriver.exe")
    driver = webdriver.Chrome(chromedriver_path)
else:
    driver = webdriver.Chrome()

url = 'https://portal.sungshin.ac.kr/sso/login.jsp'
driver.get(url)
action = ActionChains(driver)
driver.implicitly_wait(10)
driver.find_element_by_css_selector('#loginId_mobile').click()

action.send_keys('2970021').key_down(Keys.TAB).send_keys('lchlshp12*').key_down(Keys.TAB).key_down(Keys.TAB).key_down(Keys.ENTER).perform()
driver.implicitly_wait(10)
time.sleep(1)
driver.get('https://tis.sungshin.ac.kr/comm/nxui/staff/sso.do?menuUid=PORTAL_3201&connectDiv=1')
driver.maximize_window()



999 LHB 중고나라 검색 후 상세리스트(가격) 가져오기 2개의 driver로 coding
from selenium import webdriver
import requests, re
from bs4 import BeautifulSoup
from selenium.common.exceptions import NoSuchElementException, StaleElementReferenceException
import time
from openpyxl import Workbook, load_workbook

wb = Workbook()
sheet1 = wb.active
sheet1.title = '중고나라 상품'  # 시트명
sheet1.cell(row=1, column=2).value = '제목'
sheet1.cell(row=1, column=3).value = '작성자'
sheet1.cell(row=1, column=4).value = '날짜'
sheet1.cell(row=1, column=5).value = '가격'

driver = webdriver.Chrome('./chromedriver.exe')
driver2 = webdriver.Chrome('./chromedriver.exe')

joonggonara_url = 'https://cafe.naver.com/joonggonara.cafe?iframe_url=/ArticleList.nhn%3Fsearch.clubid=10050146%26search.boardtype=L%26viewType=pc'

driver.get(joonggonara_url)

# keyword = input("키워드 입력 : ")
search_input = driver.find_element_by_css_selector('input#topLayerQueryInput')
search_input.send_keys('갤럭시 탭')
# search_input.send_keys(keyword)
search_button = driver.find_element_by_css_selector("form[name='frmBoardSearch'] > button")
search_button.click()

time.sleep(1)
driver.switch_to.frame("cafe_main")

row = 2

for page in range(1, 2):
    list = driver.find_elements_by_css_selector('#main-area > div:nth-child(7) > table > tbody > tr')
    for item in list:
        title = item.find_element_by_css_selector('a.article').text.strip()
        title = re.sub('[^0-9a-zA-Zㄱ-힗]', '', title)
        writer = item.find_element_by_css_selector('a.m-tcol-c').text.strip()
        ddate = item.find_element_by_css_selector('td.td_date').text.strip()
        link = item.find_element_by_css_selector('a.article').get_attribute('href')

        driver2.get(link)
        time.sleep(1)
        driver2.switch_to.frame("cafe_main")
        try:
            cost = driver2.find_element_by_css_selector('span.cost').text
        except NoSuchElementException:
            cost = 'X'

        sheet1.cell(row=row, column=2).value = title
        sheet1.cell(row=row, column=2).hyperlink = link
        sheet1.cell(row=row, column=3).value = writer
        sheet1.cell(row=row, column=4).value = ddate
        sheet1.cell(row=row, column=5).value = cost

        row = row + 1

    if page % 10 == 0:
        driver.find_element_by_link_text('다음').click()
    else:
        next = str(page + 1)
        driver.find_element_by_link_text(next).click()

wb.save("./test4.xlsx")

item.find_element_by_css_selector('a').click()  #가격을 구하기 위해서



 999 중고나라, 웹크롤링, headless,  키워드 입력, 제외할 문자
from selenium import webdriver
from bs4 import BeautifulSoup
import datetime
import time
import smtplib
from email.mime.text import MIMEText
import os
from os.path import join, dirname
from dotenv import load_dotenv

# 고정 변수

utc_time = datetime.datetime.utcnow()
time_gap = datetime.timedelta(hours=9)
kor_time = utc_time + time_gap
today = kor_time.strftime("%Y.%m.%d.")


naver_login_url = 'https://nid.naver.com/nidlogin.login'
joonggonara_url = 'https://cafe.naver.com/joonggonara.cafe?iframe_url=/ArticleList.nhn%3Fsearch.clubid=10050146%26search.boardtype=L%26viewType=pc'
keyword_list = [
    '소니 a5100'
]
result_dic = {}
for keyword in keyword_list:
    result_dic[keyword] = {
        'title': [],
        'writer': [],
        'href': [],
        'date': []
    }
exception_flag = 0
exception_title_keyword_list = ['삽니다', '사기', '부산', '대전', '대구', '사 기', '사  기', 'ㅅ ㅏㄱ ㅣ', '완료', '경남', '창원']
exception_writer_keyword_list = []

# Headless chrome 사용법에 대해서는 아래 URL을 참고한다.
# https://beomi.github.io/gb-crawling/posts/2017-09-28-HowToMakeWebCrawler-Headless-Chrome.html

options = webdriver.ChromeOptions()
options.add_argument('headless')
options.add_argument('window-size=1920x1080')
options.add_argument("disable-gpu")

# driver = webdriver.Chrome('./chromedriver', options=options)
driver = webdriver.Chrome('./chromedriver')

# driver.get('http://naver.com')
# driver.implicitly_wait(3)
# driver.get_screenshot_as_file('naver_main_headless.png')

# 네이버 로그인
# 생략

# 중고나라 접속 및 검색어 크롤링
driver.get(joonggonara_url)
driver.implicitly_wait(3)
# driver.get_screenshot_as_file('naver_main_headless.png')

search_input = driver.find_element_by_css_selector('input#topLayerQueryInput')
for keyword in keyword_list:
    search_input.send_keys(keyword)
    search_button = driver.find_element_by_css_selector("form[name='frmBoardSearch'] > button")
    search_button.click()
    driver.implicitly_wait(3)
    time.sleep(2) # 이것을 안해줄시는 page source가 바로 전 page를 끌어와서 진행됨.
    # driver.get_screenshot_as_file('naver_main_headless.png')
    iframe = driver.find_element_by_css_selector('iframe#cafe_main')
    driver.switch_to.frame(iframe)

    # show_element = driver.find_element_by_xpath("""//*[@id="listSizeSelectDiv"]/a""")
    # show_element.click()
    # show_50_element = driver.find_element_by_xpath("""//*[@id="listSizeSelectDiv"]/ul/li[7]/a""")
    # show_50_element.click()

    req = driver.page_source
    html = BeautifulSoup(req, 'html.parser')
    print(html)
    title_list = []
    writer_list = []
    href_list = []
    date_list = []
    # driver.get_screenshot_as_file('naver_main_headless.png')
    for tag in html.select('div#content-area div#main-area table tbody tr'):
        if len(tag.select('div.inner_list > a.article')) < 1:
            continue

        title = tag.select('div.inner_list > a.article')[0].text.strip()
        print(title)
        number = tag.select('div.inner_number')[0].text.strip()
        writer = tag.select('td.p-nick > a.m-tcol-c')[0].text.strip()
        date = tag.select('td.td_date')[0].text.strip()

        if ':' in date:
            date = today

        # 제목 예외처리
        for exception_title_keyword in exception_title_keyword_list:
            if exception_title_keyword in title:
                exception_flag = 1
                break
        # 글쓴이 예외처리
        for exception_writer_keyword in exception_writer_keyword_list:
            if exception_writer_keyword == writer:
                exception_flag = 1
                break

        if exception_flag == 1:
            exception_flag = 0
            continue

        href = 'https://cafe.naver.com/joonggonara/' + number
        # print(title,"//",writer,"//",date)
        # print(href)

        if writer in writer_list:
            pass
        else:
            title_list.append(title)
            writer_list.append(writer)
            href_list.append(href)
            date_list.append(date)
    result_dic[keyword]['title'] = title_list
    result_dic[keyword]['writer'] = writer_list
    result_dic[keyword]['href'] = href_list
    result_dic[keyword]['date'] = date_list
    driver.switch_to.default_content()
# driver.quit()
# print(result_dic)

# # 파일로 기록
#
#
# # 1주일 이전 파일 삭제
#
#
# # 메일 발송
# # 메일 본문 작성 먼저
# mail_html = "<html><head></head><body>" + kor_time.strftime(
#     "%Y/%m/%d %H:%M:%S") + "<br>"  # YYYY/mm/dd HH:MM:SS 형태의 시간 출력
# for keyword in keyword_list:
#     mail_html += "<h1>" + keyword + " 크롤링 결과</h1>"
#     for i, v in enumerate(result_dic[keyword]['title']):
#         mail_html += "<p><a href='" + result_dic[keyword]['href'][i] + "'>" + v + " _ " + result_dic[keyword]['writer'][
#             i] + "</a></p>"
#         mail_html += "<p>" + result_dic[keyword]['date'][i] + "</p><br>"
#     mail_html += "<br>"
# mail_html += "</body></html>"
# print(mail_html)

# # Create .env file path.
# dotenv_path = join(dirname(__file__), '.env')
# # Load file from the path.
# load_dotenv(dotenv_path)
# # open SMTP
# smtp = smtplib.SMTP('smtp.gmail.com', 587)
# smtp.ehlo()  # say Hello
# smtp.starttls()  # TLS 사용시 필요
# smtp.login('doobw@likelion.org', os.getenv('PASSWORD'))
#
# # main html
# msg = MIMEText(mail_html, 'html')
# # title
# msg['Subject'] = '중고나라 크롤링 결과 보고'
# # from
# msg['From'] = os.getenv("FROM")
# # to
# msg['To'] = os.getenv("TO")
# # from / to / msg
# smtp.sendmail(msg['From'], msg['To'].split(','), msg.as_string())
#
# smtp.quit()
#
# # © 2020
# # GitHub, Inc.
# # Terms
# # Privacy
# # Security
# # Status
# # Help
# # Contact
# # GitHub
# # Pricing
# # API
# # Training
# # Blog
# # About



# 999 여러페이지 클릭 크롤링 중고나라에서 크롤링으로 데이터 뽑기
from selenium import webdriver
import time

keyword = input(" 키워드 입력 : ")
driver = webdriver.Chrome('./chromedriver')
driver.get(
    'https://cafe.naver.com/joonggonara?iframe_url=/ArticleList.nhn%3Fsearch.clubid=10050146%26search.menuid=1900%26search.boardtype=L')
bb2 = []
dataset = []
time.sleep(2)
search_input = driver.find_element_by_css_selector('input#topLayerQueryInput')
search_input.send_keys(keyword)

time.sleep(1)
driver.switch_to.frame("cafe_main")
for l in range(1, 71):
    tag = driver.find_elements_by_xpath('//div[@class="article-board m-tcol-c"]//table/tbody/tr')
    bb2 = tag
    time.sleep(2)

    for i in range(len(bb2)):
        dataset.append(bb2[i].text)

    if l % 10 == 0:
        c10 = driver.find_element_by_link_text('다음')
        c10.click()
    else:
        a = str(l + 1)
        c = driver.find_element_by_link_text(a)
        c.click()
    time.sleep(2)

driver.quit()
print(dataset[:20])


# 999 엑셀읽기 - 가수명 노래명으로 구글검색 - 해당하는 결과값이 있는지 점검 - 해당결과값 클릭하여 상세정보 읽기
import time
from selenium import webdriver
# import xlrd
from openpyxl import Workbook, load_workbook
from xlutils.copy import copy
from selenium.common.exceptions import NoSuchElementException, StaleElementReferenceException
import os

# 엑셀 파일 열기
wb = load_workbook("test2.xlsx")
sheet = wb.active

# driver로 wiki 접속
driver = webdriver.Chrome("./chromedriver.exe")

for i in range(2, sheet.max_row+1): # 1행은 구분자 2행부터 data
    driver.get("https://google.com")

    # 검색 키워드 셀에서 가져오기
    song = sheet.cell(row=i, column=2).value
    artist = sheet.cell(row=i, column=3).value
    keyword = str(song) + ' ' + str(artist)  # 숫자인 경우가 있어서 str()

    # 구글 검색
    elem = driver.find_element_by_xpath('//*[@id="tsf"]/div[2]/div[1]/div[1]/div/div[2]/input')

    # 공통
    elem.send_keys(keyword)
    elem.submit()

    ##### 구글 검색 결과 페이지 #####
    try:
        #box = driver.find_element_by_xpath("//*[@id='rso']")
        # list = box.find_elements_by_tag_name('h3')
        list = driver.find_elements_by_css_selector('div.g')
        for item in list:
            # print(item.text)
            if ('Wikipedia' in item.text):
                p = item.find_element_by_tag_name('a')
                p.click()
                break


    except NoSuchElementException:

        box = driver.find_element_by_xpath("//div[@id='rso']/div/div")
        # list = box.find_elements_by_tag_name('h3')
        list = box.find_elements_by_css_selector('div.g')
        for item in list:
            # print(item.text)
            if ('Wikipedia' in item.text):
                p = item.find_element_by_tag_name('a')
                p.click()
                break

    ##### 위키 곡정보 페이지  #####
    # writer,producer
    print("//////////////////////////// " + str(i) + "행을 크롤링중입니다.")
    print("//////////////////////////// " + "노래명, 가수 : " + keyword)

    try:
        table = driver.find_element_by_tag_name('table')
        tbody = table.find_element_by_tag_name("tbody")
        trs = tbody.find_elements_by_tag_name("tr")

    except NoSuchElementException:
        print(" [예외 발생] 표 없음 ")
        continue

    for tr in trs:

        if "Songwriter" in tr.text:
            print(tr.text)
            a = ""
            if tr.find_elements_by_tag_name("li"):
                lis = tr.find_elements_by_tag_name("li")
                for li in lis:
                    a = a + "," + li.text
            else:
                o = tr.find_elements_by_tag_name("td")
                a = a + "," + o[0].text

            a = a[1:]
            sheet.cell(i, 4).value = a

        if "Producer" in tr.text:
            print(tr.text)
            a = ""
            if tr.find_elements_by_tag_name("li"):
                lis = tr.find_elements_by_tag_name("li")
                for li in lis:
                    a = a + "," + li.text
            else:
                o = tr.find_elements_by_tag_name("td")
                a = a + "," + o[0].text

            a = a[1:]
            sheet.cell(i, 5).value = a

# 저장
wb.save('test2.xlsx')
view
rawwiki_crawler
hosted
with ❤ by GitHub


999 웹상의 반복되는 찾고자하는 요소들을 찾아 핸들링 색 찾기 게임 [btn.value_of_css_property('background-color') for btn in btns]
from selenium import webdriver
from pprint import  pprint
import time
from collections import Counter

driver = webdriver.Chrome('chromedriver.exe')
driver.get('http://zzzscore.com/color/')
driver.implicitly_wait(300)

btns = driver.find_elements_by_xpath('//*[@id="grid"]/div')

def analysis():
    btns_rgba = [btn.value_of_css_property('background-color') for btn in btns]
    #pprint(btns_rgba)
    result = Counter(btns_rgba)
    #print(result)
    for key, value in result.items():
        if value == 1:
            answer = key
            break
    else:
        answer = None
        print("정답을 찾을 수 없습니다.")

    if answer :
        index = btns_rgba.index(answer)
        btns[index].click()

if __name__ == "__main__":
    start = time.time()
    while time.time() - start <= 60:
        analysis()


유튜브 동영상 다운받기
from pytube import YouTube
yt = YouTube('https://www.youtube.com/watch?v=C0xIwnh8fOs') # 유튜브 영상 URL 입력
yt.streams.filter(progressive=True, file_extension='mp4').order_by('resolution').desc().first().download()



999 웹상의 반복되는 찾고자하는 요소들을 찾아 핸들링 1to50게임 숫자 빨리 찾기 find_elements_by_xpath
from selenium import webdriver
import time

driver = webdriver.Chrome('chromedriver.exe')
driver.get('https://zzzscore.com/1to50')
driver.implicitly_wait(300)
num = 1

btns = driver.find_elements_by_xpath('//*[@id="grid"]/div[*]')

for i in range(0, 50):
    for j in range(0, len(btns)):
        print('j =  ', j)
        if num == int(btns[j].text):
            btn = btns[j]
            btn.click()

            time.sleep(0.1)
            # print(num, ' ', j)
            num = num + 1
            break
    btns = driver.find_elements_by_xpath('//*[@id="grid"]/div[*]')
    print(num, ' ', j, '   ', btns[j-1].text)

전체에서 필요한 list를 얻고 list를 가지고 하나씩 뽑아서 분석
import requests
from bs4 import BeautifulSoup

req = requests.get('http://tv.naver.com/r')
raw = req.text
html = BeautifulSoup(raw, 'html.parser')

infos = html.select('div.cds_area > div')
print(infos[0])
for info in infos:
    title = info.select_one('dt.title > a').text
    chn = info.select_one('dd.chn').text[0:-1]
    hit = info.select_one('span.hit').text
    like = info.select_one('span.like').text

    print(chn, '/', title, '/', hit, '/', like)

999 전체에서 통으로 원하는data를 얻고 거기서 list를 받아서 하나씩 분석 / 특수문자처리 / 이미지 웹에서 끌어와 저장 / 이미지 폴더에서 끌어오기 이미지 엑셀에 저장
from bs4 import BeautifulSoup as bs
from pprint import pprint
import requests, re, os
from urllib.request import  urlretrieve
from openpyxl.drawing.image import Image
from openpyxl import Workbook

wb = Workbook()
sheet1 = wb.active
sheet1.title = '네이버 웹툰 완결편'

try:
    if not (os.path.isdir('image')):
        os.mkdir(os.path.join('image'))
except OSError as e:
    if e.errno != errno.EEXITS:
        print("폴더생성실패")
        exit()
url = 'https://comic.naver.com/webtoon/finish.nhn'
html = requests.get(url)
soup = bs(html.text, 'html.parser')
all_data = soup.find('div', {'class':'list_area'})

data_list = all_data.findAll('li')

col = 1
row = 1
for data in data_list:
    img_data = data.find('img')
    img_src = img_data['src']
    a_list = data.find('a')
    title = a_list['title']
    title = re.sub('[^0-9a-zA-Zㄱ-힗]', '', title)
    link = "https//commic.naver.com" + a_list['href']
    strong = data.find('strong').text

#   urlretrieve(img_src, './image/'+title+'.gif')
    img_title = './image/'+title+'.gif'
    img_file = Image(img_title)
    print(img_title)
    img_file.anchor= 'C10'
    #cell = sheet1.cell(row=10, column=1)
    sheet1.add_image(img_file)
#    sheet1.cell()
#    col = col + 1
    row = row + 1
    break
wb.save("./webtoon.xlsx")


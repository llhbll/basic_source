from openpyxl import Workbook, load_workbook

wb = load_workbook("./data.xlsx")
sheet = wb.active
sheet2 = wb.create_sheet('data_sheet', 2)

# wb1 = Workbook()
# sheet1 = wb1.active

except_str = ['전문교육', '2018', '2020', '2학기', '특근식대', '부서운영비', '바우처']
inclu_str = ['학점은행', '2019', '1학기']
data_row = 1
for row in range(2, sheet.max_row + 1):
    title = sheet.cell(row=row, column=8).value
    gubun = sheet.cell(row=row, column=5).value
    cost = sheet.cell(row=row, column=9).value

    exception_flag = 0
    for out_word in except_str:
        if out_word in title:
            exception_flag = 1
            break

    if exception_flag == 0:
        data_row = data_row + 1
        sheet2.cell(row=data_row, column= 1).value = title
        sheet2.cell(row=data_row, column= 2).value = cost
        sheet2.cell(row=data_row, column= 3).value = gubun

wb.save("./data.xlsx")
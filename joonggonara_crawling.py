from selenium import webdriver
import requests, re
from bs4 import BeautifulSoup
from selenium.common.exceptions import NoSuchElementException, StaleElementReferenceException
import time
from openpyxl import Workbook, load_workbook
import pyperclip
from selenium.webdriver.common.keys import Keys
import sys

wb  = Workbook()
sheet1 = wb.active
sheet1.title = '중고나라 상품' #시트명
sheet1.cell(row=1, column=2).value = '제목'
sheet1.cell(row=1, column=3).value = '작성자'
sheet1.cell(row=1, column=4).value = '날짜'
sheet1.cell(row=1, column=5).value = '가격'

def uf_get_idpw(siteName):
    f = open("./pw.txt", 'r') # 실행할 폴더(바탕화면이면 바탕화면에)에 pw.txt file이 있어야 함.
    for data in f.readlines():
        site = data.split(':')[0]
        id = data.split(':')[1]
        pw = data.split(':')[2].strip('\n')
        pw = 'lch' + pw #pw file이 노출되었을때를 대비하여 앞 3글자는 여기서
        if site == siteName:
            return id, pw
keyword = input("검색 키워드 입력 : ")
page_cnt = input("검색할 page수 : ")
#
if getattr(sys, 'frozen', False):
    chromedriver_path = os.path.join(sys._MEIPASS, "chromedriver.exe")
    driver = webdriver.Chrome(chromedriver_path)
    driver2 = webdriver.Chrome(chromedriver_path)
else:
    driver = webdriver.Chrome()
    driver2 = webdriver.Chrome()

# driver = webdriver.Chrome('./chromedriver.exe')
# driver2 = webdriver.Chrome('./chromedriver.exe')
#
#
# 자동입력 방지를 뚫기위해 pyperclip을 사용 로긴
user_id, user_pw = uf_get_idpw('naver')
naverlogin_url = 'https://nid.naver.com/nidlogin.login?mode=form&url=https%3A%2F%2Fwww.naver.com'
driver.get(naverlogin_url)
time.sleep(1)
naver_id_input = driver.find_element_by_css_selector('#id')
naver_id_input.clear()
time.sleep(1)

naver_id_input.click()
pyperclip.copy(user_id)
naver_id_input.send_keys(Keys.CONTROL, 'v')
time.sleep(1)

naver_pw_input = driver.find_element_by_css_selector('#pw')
naver_pw_input.click()
pyperclip.copy(user_pw)
naver_pw_input.send_keys(Keys.CONTROL, 'v')
time.sleep(1)

search_button = driver.find_element_by_css_selector("#log\.login")
search_button.click()
time.sleep(3)

joonggonara_url = 'https://cafe.naver.com/joonggonara'
# joonggonara_url = 'https://cafe.naver.com/joonggonara.cafe?iframe_url=/ArticleList.nhn%3Fsearch.clubid=10050146%26search.boardtype=L%26viewType=pc'

driver.get(joonggonara_url)

search_input = driver.find_element_by_css_selector('input#topLayerQueryInput')
# search_input.send_keys('갤럭시 탭')
search_input.send_keys(keyword)
search_button = driver.find_element_by_css_selector("form[name='frmBoardSearch'] > button")
search_button.click()

time.sleep(1)
driver.switch_to.frame("cafe_main")

except_str = [ "중나협력", "삽니다", "매입", "구입", "사요", "구매" ]
row = 2

for page in range(1, int(page_cnt) + 1):
    list = driver.find_elements_by_css_selector('#main-area > div:nth-child(7) > table > tbody > tr')
    for item in list:
        title = item.find_element_by_css_selector('a.article').text.strip()
        # title = re.sub('[^0-9a-zA-Zㄱ-힗]', '', title)
        if any(format in title for format in except_str):
            continue
        strip_title = str(title).replace(' ', '')
        aa = title.find(keyword)
        if strip_title.find(keyword.replace(' ', '')) == -1: #양쪽다 공백없이 키워드 비교
            continue
        writer = item.find_element_by_css_selector('a.m-tcol-c').text.strip()
        ddate = item.find_element_by_css_selector('td.td_date').text.strip()
        link = item.find_element_by_css_selector('a.article').get_attribute('href')

        time.sleep(2)
        driver2.get(link, auth=(user_id, user_pw)) #로긴하고 상세페이지 들어가면 모두 접근가능할 줄 알았는데 중고나라회원에 허용한 page는 접근이 안됨! 추후 방법을 찾아야 함. 현재 실력으로는 불가
        time.sleep(2)
        try:
           driver2.switch_to.frame("cafe_main")
           cost = driver2.find_element_by_css_selector('span.cost').text
        except NoSuchElementException:
            cost = 'X'

        sheet1.cell(row=row, column=2).value = title
        sheet1.cell(row=row, column=2).hyperlink = link
        sheet1.cell(row=row, column=3).value = writer
        sheet1.cell(row=row, column=4).value = ddate
        sheet1.cell(row=row, column=5).value = cost

        row = row + 1


    if page % 10 == 0:
        driver.find_element_by_link_text('다음').click()
    else:
        next = str(page + 1)
        driver.find_element_by_link_text(next).click()
# driver.quit()
# driver2.quit()
wb.save(keyword + ".xlsx")

        #item.find_element_by_css_selector('a').click()  #가격을 구하기 위해서






